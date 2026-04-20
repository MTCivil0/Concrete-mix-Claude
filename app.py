"""
app.py — Concrete Mix Design Copilot (Graduate Research Edition)
Teymouri Research Lab | South Dakota State University
Run with: streamlit run app.py
"""

import streamlit as st
import pandas as pd
import anthropic
import base64
import os
import json
import io
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.aci211 import calculate_mix, SHAPE_WATER_REDUCTION
from src.claude_client import run_analysis
from src.schemas import MixDesignInput, MixDesignResult, ProjectInfo
from src.reporting import generate_pdf_report

load_dotenv()

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG & CSS
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Concrete Mix Design Copilot | Teymouri Research Lab",
    page_icon="🧱",
    layout="wide",
)

st.markdown("""
<style>
.sdsu-header {
    background: #0033A0; color: white;
    padding: 12px 20px; border-radius: 8px 8px 0 0;
    font-weight: 700; font-size: 15px;
}
.sdsu-gold-bar { height: 4px; background: #FFB71B; border-radius: 0 0 4px 4px; margin-bottom: 16px; }
.sdsu-section {
    background: #0033A0; color: white;
    padding: 6px 14px; border-radius: 5px;
    font-size: 13px; font-weight: 600; margin: 14px 0 6px 0;
}
.sdsu-q-label {
    background: #E6EBF5; border-left: 4px solid #0033A0;
    padding: 7px 14px; border-radius: 0 6px 6px 0; margin-bottom: 8px;
    color: #0033A0; font-weight: 600; font-size: 14px;
}
.saved-badge {
    background: #E8F5E9; border: 1px solid #81C784;
    padding: 4px 10px; border-radius: 99px;
    color: #2E7D32; font-size: 12px; display: inline-block;
}
div[data-testid="column"] button {
    border-left: 3px solid #0033A0 !important; font-size: 12px !important;
}
div[data-testid="column"] button:hover {
    border-left: 3px solid #FFB71B !important; background: #E6EBF5 !important;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("**Teymouri Research Lab**")
    st.caption("Dept. of Construction & Concrete Industry Management\nJerome J. Lohr College of Engineering\nSouth Dakota State University")
    st.divider()
    demo_mode = st.toggle("Demo mode (no API cost)", value=True)
    st.caption("Turn off + add API key to .env for live Claude inference.")
    st.divider()

    # ── Saved sessions ────────────────────────────────────────────────────────
    st.markdown("**💾 Saved mix designs**")
    if "saved_mixes" not in st.session_state:
        st.session_state.saved_mixes = {}

    if st.session_state.saved_mixes:
        for name in list(st.session_state.saved_mixes.keys()):
            col_l, col_r = st.columns([3,1])
            with col_l:
                if st.button(f"📂 {name}", key=f"load_{name}", use_container_width=True):
                    st.session_state.load_mix = name
                    st.rerun()
            with col_r:
                if st.button("✕", key=f"del_{name}"):
                    del st.session_state.saved_mixes[name]
                    st.rerun()
    else:
        st.caption("No saved mixes yet. Design a mix and click Save.")

    st.divider()
    st.markdown("**📊 Mix comparison**")
    if "comparison_mixes" not in st.session_state:
        st.session_state.comparison_mixes = []
    n = len(st.session_state.comparison_mixes)
    st.caption(f"{n} mix{'es' if n != 1 else ''} in comparison table")
    if n > 0 and st.button("🗑️ Clear comparison", use_container_width=True):
        st.session_state.comparison_mixes = []
        st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="sdsu-header">🧱 &nbsp; Teymouri Research Lab · South Dakota State University</div>
<div class="sdsu-gold-bar"></div>
""", unsafe_allow_html=True)
st.title("Concrete Mix Design Copilot")
st.caption("Graduate Research Edition · ACI 211.1 · Teymouri Research Lab · SDSU")

# ─────────────────────────────────────────────────────────────────────────────
# MODE SELECTION
# ─────────────────────────────────────────────────────────────────────────────
mode = st.selectbox(
    "What would you like to do?",
    ["— select an option —",
     "🔬 Design a new concrete mix (ACI 211.1)",
     "📊 Compare saved mix designs",
     "📄 Review an existing mix design (upload file)",
     "💬 Ask a concrete question (student Q&A)"],
    index=0,
)

if mode == "— select an option —":
    st.stop()

st.divider()

# ─────────────────────────────────────────────────────────────────────────────
# HELPER: Excel export
# ─────────────────────────────────────────────────────────────────────────────
def make_excel(mixes: list) -> bytes:
    """Generate an Excel workbook from a list of (label, aci_result, inp) tuples."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mix Design Comparison"

    BLUE  = "0033A0"
    GOLD  = "FFB71B"
    LGRAY = "F5F5F5"

    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", fgColor=BLUE)
    gold_fill = PatternFill("solid", fgColor=GOLD)
    gray_fill = PatternFill("solid", fgColor=LGRAY)
    bold_font = Font(name="Arial", bold=True, size=10)
    reg_font  = Font(name="Arial", size=10)
    center    = Alignment(horizontal="center", vertical="center")
    left      = Alignment(horizontal="left", vertical="center")
    thin      = Side(style="thin", color="BDBDBD")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "Teymouri Research Lab — Concrete Mix Design Comparison"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color=BLUE)
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y %H:%M')} | ACI 211.1 (PCA Method)"
    ws["A2"].font = Font(name="Arial", size=9, color="888888")
    ws["A2"].alignment = center

    # Column headers
    row = 4
    params = [
        ("Parameter", "A"),
    ]
    ws["A4"] = "Parameter"
    ws["A4"].font = hdr_font
    ws["A4"].fill = hdr_fill
    ws["A4"].alignment = left
    ws["A4"].border = border
    ws.column_dimensions["A"].width = 30

    for i, (label, aci, inp) in enumerate(mixes):
        col = get_column_letter(i + 2)
        ws[f"{col}4"] = label
        ws[f"{col}4"].font = hdr_font
        ws[f"{col}4"].fill = hdr_fill
        ws[f"{col}4"].alignment = center
        ws[f"{col}4"].border = border
        ws.column_dimensions[col].width = 18

    def write_section(title, items, start_row, odd_fill=None):
        r = start_row
        # Section header
        ws.merge_cells(f"A{r}:{get_column_letter(len(mixes)+1)}{r}")
        ws[f"A{r}"] = title
        ws[f"A{r}"].font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        ws[f"A{r}"].fill = PatternFill("solid", fgColor="0033A0")
        ws[f"A{r}"].alignment = left
        ws[f"A{r}"].border = border
        ws.row_dimensions[r].height = 18
        r += 1

        for j, (param, vals) in enumerate(items):
            fill = gray_fill if j % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            ws[f"A{r}"] = param
            ws[f"A{r}"].font = bold_font
            ws[f"A{r}"].fill = fill
            ws[f"A{r}"].border = border
            ws[f"A{r}"].alignment = left
            for k, val in enumerate(vals):
                col = get_column_letter(k + 2)
                ws[f"{col}{r}"] = val
                ws[f"{col}{r}"].font = reg_font
                ws[f"{col}{r}"].fill = fill
                ws[f"{col}{r}"].border = border
                ws[f"{col}{r}"].alignment = center
            r += 1
        return r

    # ── Inputs section ────────────────────────────────────────────────────────
    input_rows = [
        ("Required f'c (psi)",       [f"{m[2].fc_psi:,}"           for m in mixes]),
        ("Exposure codes",            [", ".join(m[2].exposure_codes) for m in mixes]),
        ("Max agg size (in)",         [m[2].agg_size                 for m in mixes]),
        ("Target slump (in)",         [m[2].slump                    for m in mixes]),
        ("Fineness modulus (FA)",     [f"{m[2].fm_fa:.2f}"           for m in mixes]),
        ("SG coarse aggregate",       [f"{m[2].sg_ca:.2f}"           for m in mixes]),
        ("SG fine aggregate",         [f"{m[2].sg_fa:.2f}"           for m in mixes]),
        ("Air entrained",             ["Yes" if m[2].air_entrained else "No" for m in mixes]),
        ("Fly ash (%)",               [f"{m[2].flyash_pct:.0f}"      for m in mixes]),
        ("Slag (%)",                  [f"{m[2].slag_pct:.0f}"        for m in mixes]),
        ("Silica fume (%)",           [f"{m[2].sf_pct:.0f}"          for m in mixes]),
        ("PCC (%)",                   [f"{m[2].pcc_pct:.0f}"         for m in mixes]),
    ]
    r = write_section("MIX DESIGN INPUTS", input_rows, 6)

    # ── ACI results section ───────────────────────────────────────────────────
    result_rows = [
        ("Selected w/cm",             [str(m[1]["selected_wcm"])     for m in mixes]),
        ("Air content (%)",           [str(m[1]["air_pct"])          for m in mixes]),
        ("Fresh density (pcf)",       [str(m[1]["density_pcf"])      for m in mixes]),
        ("Total volume (ft³/CY)",     [str(m[1]["volumes_ft3"]["total"]) for m in mixes]),
        ("Water (lbs/CY)",            [str(m[1]["proportions"]["water_lbs"]) for m in mixes]),
        ("Portland cement (lbs/CY)",  [str(m[1]["proportions"]["cement_lbs"]) for m in mixes]),
        ("Fly ash (lbs/CY)",          [str(m[1]["proportions"]["flyash_lbs"]) for m in mixes]),
        ("Slag (lbs/CY)",             [str(m[1]["proportions"]["slag_lbs"]) for m in mixes]),
        ("Silica fume (lbs/CY)",      [str(m[1]["proportions"]["sf_lbs"]) for m in mixes]),
        ("PCC (lbs/CY)",              [str(m[1]["proportions"]["pcc_lbs"]) for m in mixes]),
        ("Coarse aggregate (lbs/CY)", [str(m[1]["proportions"]["ca_lbs"]) for m in mixes]),
        ("Fine aggregate (lbs/CY)",   [str(m[1]["proportions"]["fa_lbs"]) for m in mixes]),
        ("Total cementitious (lbs/CY)",[str(m[1]["proportions"]["total_cm_lbs"]) for m in mixes]),
    ]
    r = write_section("ACI 211.1 RESULTS", result_rows, r + 1)

    # Footer
    r += 1
    ws.merge_cells(f"A{r}:{get_column_letter(len(mixes)+1)}{r}")
    ws[f"A{r}"] = "Teymouri Research Lab · SDSU · ACI 211.1 (PCA Method) · Not a substitute for a licensed engineer"
    ws[f"A{r}"].font = Font(name="Arial", size=8, color="888888", italic=True)
    ws[f"A{r}"].alignment = center

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# MODE: DESIGN A NEW MIX
# ─────────────────────────────────────────────────────────────────────────────
if "Design" in mode:

    # Check if loading a saved mix
    load_defaults = {}
    if "load_mix" in st.session_state:
        name = st.session_state.pop("load_mix")
        if name in st.session_state.saved_mixes:
            load_defaults = st.session_state.saved_mixes[name]
            st.success(f"Loaded: **{name}**")

    def D(key, fallback):
        return load_defaults.get(key, fallback)

    # ── Project info ──────────────────────────────────────────────────────────
    with st.expander("📋 Project information (optional — appears in report)"):
        pi1, pi2 = st.columns(2)
        with pi1:
            proj_name   = st.text_input("Project name",   value=D("proj_name",""),   placeholder="e.g. US-14 Bridge Deck")
            location    = st.text_input("Location",        value=D("location",""),    placeholder="e.g. Brookings, SD")
        with pi2:
            prepared_by = st.text_input("Prepared by",    value=D("prepared_by",""), placeholder="e.g. Dr. Mohammad Teymouri, PE")
            company     = st.text_input("Organization",   value=D("company",""),     placeholder="e.g. SDSU Teymouri Research Lab")

    with st.expander("🏭 Material producers (optional)"):
        mp1, mp2 = st.columns(2)
        with mp1:
            cement_prod = st.text_input("Cement producer", value=D("cement_prod",""), placeholder="e.g. Ash Grove, Type I/II")
            flyash_prod = st.text_input("Fly ash producer", value=D("flyash_prod",""), placeholder="e.g. Basin Electric, Class F")
            slag_prod   = st.text_input("Slag producer",   value=D("slag_prod",""),   placeholder="e.g. Lafarge, Grade 100")
        with mp2:
            sf_prod     = st.text_input("Silica fume",     value=D("sf_prod",""),     placeholder="e.g. Elkem, densified")
            pcc_prod    = st.text_input("PCC producer",    value=D("pcc_prod",""),    placeholder="e.g. Western Sugar Cooperative")

    st.divider()

    # ── Mix inputs ────────────────────────────────────────────────────────────
    with st.form("mix_form"):
        st.markdown('<div class="sdsu-section">I. Exposure & strength</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            fc_psi        = st.number_input("Required f'c (psi)", 2000, 12000, D("fc_psi",5000), 500)
            ft_class      = st.selectbox("Freeze-thaw", ["F0 — protected","F1 — limited moisture","F2 — moisture exposed","F3 — deicers"], index=D("ft_idx",1))
            water_class   = st.selectbox("Water exposure", ["W0 — protected","W1 — low permeability required"], index=D("w_idx",0))
        with c2:
            sulfate_class = st.selectbox("Sulfate", ["S0 — protected","S1 — moderate","S2 — severe","S3 — very severe"], index=D("s_idx",0))
            chloride_class= st.selectbox("Chloride", ["C0 — protected","C2 — deicers / seawater"], index=D("c_idx",0))
            air_entrained = st.checkbox("Air-entrained mix", value=D("air_entrained",True))

        st.markdown('<div class="sdsu-section">II. Aggregate & workability</div>', unsafe_allow_html=True)
        c3, c4 = st.columns(2)
        with c3:
            agg_size      = st.selectbox("Nominal max agg size", ["3/8","1/2","3/4","1","1-1/2","2"], index=D("agg_idx",2), help="1/5 narrowest dim, 3/4 rebar spacing, 1/3 slab depth")
            slump         = st.selectbox("Target slump (in)", ["1-2","3-4","6-7"], index=D("slump_idx",0))
            agg_shape     = st.selectbox("Aggregate shape", list(SHAPE_WATER_REDUCTION.keys()), index=D("shape_idx",2))
        with c4:
            fm_fa         = st.number_input("Fineness modulus (FA)", 2.3, 3.1, D("fm_fa",2.77), 0.01, help="From your sieve analysis — RSCA Excel default: 2.77")
            sg_ca         = st.number_input("SG coarse aggregate", 2.4, 3.0, D("sg_ca",2.65), 0.01)
            sg_fa         = st.number_input("SG fine aggregate", 2.4, 3.0, D("sg_fa",2.68), 0.01)
            rodded_density= st.number_input("Rodded bulk density CA (lbs/ft³)", 70.0, 120.0, D("rodded",100.0), 1.0, help="ASTM C29")

        st.markdown('<div class="sdsu-section">III. SCMs</div>', unsafe_allow_html=True)
        cs1, cs2, cs3, cs4 = st.columns(4)
        with cs1: flyash_pct = st.number_input("Fly ash (%)", 0, 40, D("flyash_pct",0), 5)
        with cs2: slag_pct   = st.number_input("Slag (%)",    0, 50, D("slag_pct",0),   5)
        with cs3: sf_pct     = st.number_input("Silica fume (%)", 0, 15, D("sf_pct",0), 1)
        with cs4: pcc_pct    = st.number_input("PCC (%)",     0, 20, D("pcc_pct",5),    1, help="Precipitated Calcium Carbonate — inert micro-filler. Optimal: 3–8%.")

        total_scm = flyash_pct + slag_pct + sf_pct + pcc_pct
        if total_scm > 0:
            st.caption(f"Total SCM: {total_scm}%  →  Portland cement: {100-total_scm}%")

        with st.expander("Advanced: SCM specific gravities"):
            sg1,sg2,sg3,sg4 = st.columns(4)
            with sg1: sg_flyash = st.number_input("SG fly ash", 2.0, 3.0, D("sg_flyash",2.65), 0.01)
            with sg2: sg_slag   = st.number_input("SG slag",    2.0, 3.0, D("sg_slag",2.85),   0.01)
            with sg3: sg_sf     = st.number_input("SG sil. fume", 2.0, 2.5, D("sg_sf",2.20),   0.01)
            with sg4: sg_pcc    = st.number_input("SG PCC",     2.0, 3.0, D("sg_pcc",2.71),    0.01, help="Western Sugar PCC ≈ 2.71")

        wcm_on = st.checkbox("Override w/cm manually")
        wcm_override = st.slider("w/cm override", 0.28, 0.70, D("wcm_ov",0.45), 0.01) if wcm_on else None

        st.markdown('<div class="sdsu-section">IV. Notes</div>', unsafe_allow_html=True)
        mix_label   = st.text_input("Mix label (for comparison table)", value=D("mix_label",""), placeholder="e.g. Mix A — 5% PCC, F2 exposure")
        field_notes = st.text_area("Field notes / research context", value=D("field_notes",""), placeholder="e.g. Thesis Chapter 3 — parametric study of PCC replacement...", height=80)

        submitted = st.form_submit_button("🔬 Run Mix Design Analysis", use_container_width=True)

    # ── Parametric sensitivity slider (outside form) ──────────────────────────
    st.divider()
    st.markdown("**⚡ Parametric sensitivity — adjust PCC % and see instant recalculation**")
    st.caption("Move the slider to explore how PCC replacement affects proportions without resubmitting the form.")
    param_pcc = st.slider("PCC replacement (%)", 0, 20, int(pcc_pct), 1, key="param_pcc")

    if param_pcc != pcc_pct:
        codes_p = [ft_class.split(" ")[0], water_class.split(" ")[0],
                   sulfate_class.split(" ")[0], chloride_class.split(" ")[0]]
        aci_p = calculate_mix(
            fc_psi=fc_psi, exposure_codes=codes_p, agg_size=agg_size,
            slump=slump, fm_fa=fm_fa, sg_ca=sg_ca, sg_fa=sg_fa,
            rodded_density_ca=rodded_density, agg_shape=agg_shape,
            flyash_pct=float(flyash_pct), slag_pct=float(slag_pct),
            sf_pct=float(sf_pct), pcc_pct=float(param_pcc),
            air_entrained=air_entrained,
        )
        pp = aci_p["proportions"]
        pm1,pm2,pm3,pm4 = st.columns(4)
        pm1.metric("Water", f"{pp['water_lbs']:.0f} lbs/CY")
        pm2.metric("Cement", f"{pp['cement_lbs']:.0f} lbs/CY")
        pm3.metric("PCC", f"{pp['pcc_lbs']:.0f} lbs/CY", delta=f"PCC @ {param_pcc}%")
        pm4.metric("w/cm", f"{aci_p['selected_wcm']:.2f}")

    # ── Results ───────────────────────────────────────────────────────────────
    if submitted:
        if total_scm >= 100:
            st.error("Total SCM cannot be 100% or more.")
            st.stop()

        codes = [ft_class.split(" ")[0], water_class.split(" ")[0],
                 sulfate_class.split(" ")[0], chloride_class.split(" ")[0]]

        pi = ProjectInfo(project_name=proj_name, location=location,
                         prepared_by=prepared_by, company=company,
                         cement_producer=cement_prod, flyash_producer=flyash_prod,
                         slag_producer=slag_prod, sf_producer=sf_prod, pcc_producer=pcc_prod)

        inp = MixDesignInput(
            fc_psi=fc_psi, exposure_codes=codes, agg_size=agg_size,
            slump=slump, fm_fa=fm_fa, sg_ca=sg_ca, sg_fa=sg_fa,
            rodded_density_ca=rodded_density, agg_shape=agg_shape,
            flyash_pct=float(flyash_pct), slag_pct=float(slag_pct),
            sf_pct=float(sf_pct), pcc_pct=float(pcc_pct),
            sg_flyash=sg_flyash, sg_slag=sg_slag, sg_sf=sg_sf, sg_pcc=sg_pcc,
            air_entrained=air_entrained, wcm_override=wcm_override,
            field_notes=field_notes, project_info=pi,
        )

        with st.spinner("Running ACI 211.1 calculations + AI analysis..."):
            aci = calculate_mix(
                fc_psi=inp.fc_psi, exposure_codes=inp.exposure_codes,
                agg_size=inp.agg_size, slump=inp.slump, fm_fa=inp.fm_fa,
                sg_ca=inp.sg_ca, sg_fa=inp.sg_fa, rodded_density_ca=inp.rodded_density_ca,
                agg_shape=inp.agg_shape, flyash_pct=inp.flyash_pct, slag_pct=inp.slag_pct,
                sf_pct=inp.sf_pct, pcc_pct=inp.pcc_pct, sg_flyash=inp.sg_flyash,
                sg_slag=inp.sg_slag, sg_sf=inp.sg_sf, sg_pcc=inp.sg_pcc,
                air_entrained=inp.air_entrained, wcm_override=inp.wcm_override,
            )
            result = run_analysis(inp, aci, demo_mode)

        st.divider()
        st.markdown('<div class="sdsu-section">Results</div>', unsafe_allow_html=True)

        risk_icons = {"Low":"🟢","Moderate":"🟡","High":"🔴"}
        st.subheader(f"{risk_icons.get(result.risk_level,'🟡')} Risk: {result.risk_level}")
        st.info(result.ai_analysis)

        p = aci["proportions"]
        cols = st.columns(5)
        cols[0].metric("Design f'c",    f"{aci['design_fc_psi']:,} psi")
        cols[1].metric("w/cm",          f"{aci['selected_wcm']:.2f}")
        cols[2].metric("Air content",   f"{aci['air_pct']}%")
        cols[3].metric("Density",       f"{aci['density_pcf']} pcf")
        cols[4].metric("Volume/CY",     f"{aci['volumes_ft3']['total']:.2f} ft³")

        # Proportions table
        rows = [("Water", f"{p['water_lbs']:.0f}", "—"),
                ("Portland cement", f"{p['cement_lbs']:.0f}", f"{100-total_scm:.0f}% of CM")]
        if flyash_pct>0: rows.append(("Fly ash", f"{p['flyash_lbs']:.0f}", f"{flyash_pct:.0f}%"))
        if slag_pct>0:   rows.append(("Slag",    f"{p['slag_lbs']:.0f}",   f"{slag_pct:.0f}%"))
        if sf_pct>0:     rows.append(("Silica fume", f"{p['sf_lbs']:.0f}", f"{sf_pct:.0f}%"))
        if pcc_pct>0:    rows.append(("PCC (micro-filler)", f"{p['pcc_lbs']:.0f}", f"{pcc_pct:.0f}%"))
        rows += [("Coarse aggregate (SSD)", f"{p['ca_lbs']:.0f}", f"BV={aci['bv_ca']:.2f}"),
                 ("Fine aggregate (SSD)",   f"{p['fa_lbs']:.0f}", "abs. volume"),
                 ("Total cementitious",     f"{p['total_cm_lbs']:.0f}", "")]
        st.dataframe(pd.DataFrame(rows, columns=["Material","lbs/CY","Notes"]),
                     use_container_width=True, hide_index=True)

        for f in aci["flags"]:
            if f["status"]=="ok":      st.success(f["flag"])
            elif f["status"]=="warning": st.warning(f["flag"])
            else:                        st.error(f["flag"])

        ca, cb, cc = st.columns(3)
        with ca:
            st.markdown("**SCM & PCC notes**")
            for n in result.scm_notes: st.markdown(f"- {n}")
        with cb:
            st.markdown("**ACI 211.1 compliance**")
            for n in result.aci_compliance: st.markdown(f"- {n}")
        with cc:
            st.markdown("**QC tests**")
            for t in result.qc_tests: st.markdown(f"- {t}")

        if result.recommendations:
            st.markdown("**Recommendations**")
            st.markdown(result.recommendations)

        # ── Save & export actions ─────────────────────────────────────────────
        st.divider()
        st.markdown("**Actions**")
        act1, act2, act3, act4 = st.columns(4)

        with act1:
            save_name = st.text_input("Save as:", placeholder="e.g. Mix A — PCC 5%")
        with act2:
            st.write("")
            st.write("")
            if st.button("💾 Save mix design", use_container_width=True):
                if save_name:
                    st.session_state.saved_mixes[save_name] = {
                        "fc_psi":fc_psi, "ft_idx":["F0 — protected","F1 — limited moisture","F2 — moisture exposed","F3 — deicers"].index(ft_class),
                        "w_idx":["W0 — protected","W1 — low permeability required"].index(water_class),
                        "s_idx":["S0 — protected","S1 — moderate","S2 — severe","S3 — very severe"].index(sulfate_class),
                        "c_idx":["C0 — protected","C2 — deicers / seawater"].index(chloride_class),
                        "air_entrained":air_entrained, "agg_idx":["3/8","1/2","3/4","1","1-1/2","2"].index(agg_size),
                        "slump_idx":["1-2","3-4","6-7"].index(slump),
                        "shape_idx":list(SHAPE_WATER_REDUCTION.keys()).index(agg_shape),
                        "fm_fa":fm_fa, "sg_ca":sg_ca, "sg_fa":sg_fa, "rodded":rodded_density,
                        "flyash_pct":flyash_pct, "slag_pct":slag_pct, "sf_pct":sf_pct, "pcc_pct":pcc_pct,
                        "sg_flyash":sg_flyash, "sg_slag":sg_slag, "sg_sf":sg_sf, "sg_pcc":sg_pcc,
                        "wcm_ov":wcm_override or 0.45, "mix_label":mix_label or save_name,
                        "field_notes":field_notes, "proj_name":proj_name, "location":location,
                        "prepared_by":prepared_by, "company":company,
                        "cement_prod":cement_prod, "flyash_prod":flyash_prod,
                        "slag_prod":slag_prod, "sf_prod":sf_prod, "pcc_prod":pcc_prod,
                    }
                    st.success(f"Saved as **{save_name}**")
                    st.rerun()
                else:
                    st.warning("Enter a name first.")

        with act3:
            st.write("")
            st.write("")
            if st.button("📊 Add to comparison", use_container_width=True):
                label = mix_label or f"Mix {len(st.session_state.comparison_mixes)+1}"
                st.session_state.comparison_mixes.append((label, aci, inp))
                st.success(f"Added **{label}** to comparison — {len(st.session_state.comparison_mixes)} mix(es) total.")

        with act4:
            pdf_bytes = generate_pdf_report(result)
            fname = f"mix_{proj_name.replace(' ','_') or 'design'}.pdf"
            st.write("")
            st.download_button("📄 Download PDF report", pdf_bytes, fname, "application/pdf", use_container_width=True)


# ─────────────────────────────────────────────────────────────────────────────
# MODE: COMPARE MIX DESIGNS
# ─────────────────────────────────────────────────────────────────────────────
elif "Compare" in mode:
    st.subheader("📊 Mix Design Comparison")

    mixes = st.session_state.get("comparison_mixes", [])

    if not mixes:
        st.info("No mixes in the comparison table yet. Go to **Design a new mix**, run an analysis, and click **Add to comparison**.")
        st.stop()

    st.caption(f"{len(mixes)} mix(es) loaded. Design more mixes and add them to build your comparison.")

    # Build comparison dataframe
    params = ["Required f'c (psi)","Exposure codes","Max agg size","Slump (in)",
              "w/cm","Air content (%)","Density (pcf)","Volume (ft³/CY)",
              "Water (lbs/CY)","Portland cement (lbs/CY)","Fly ash (lbs/CY)",
              "Slag (lbs/CY)","Silica fume (lbs/CY)","PCC (lbs/CY)",
              "Coarse agg (lbs/CY)","Fine agg (lbs/CY)","Total CM (lbs/CY)"]

    data = {"Parameter": params}
    for label, aci, inp in mixes:
        p = aci["proportions"]
        data[label] = [
            f"{inp.fc_psi:,}",
            ", ".join(inp.exposure_codes),
            f'{inp.agg_size}"',
            f'{inp.slump}"',
            str(aci["selected_wcm"]),
            f"{aci['air_pct']}%",
            str(aci["density_pcf"]),
            str(aci["volumes_ft3"]["total"]),
            str(p["water_lbs"]),
            str(p["cement_lbs"]),
            str(p["flyash_lbs"]),
            str(p["slag_lbs"]),
            str(p["sf_lbs"]),
            str(p["pcc_lbs"]),
            str(p["ca_lbs"]),
            str(p["fa_lbs"]),
            str(p["total_cm_lbs"]),
        ]

    df = pd.DataFrame(data).set_index("Parameter")
    st.dataframe(df, use_container_width=True)

    st.divider()
    ex_col1, ex_col2 = st.columns(2)
    with ex_col1:
        xlsx_bytes = make_excel(mixes)
        st.download_button("📥 Download comparison as Excel (.xlsx)",
                           xlsx_bytes,
                           "mix_design_comparison.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
    with ex_col2:
        csv_data = df.to_csv().encode()
        st.download_button("📥 Download as CSV",
                           csv_data, "mix_design_comparison.csv", "text/csv",
                           use_container_width=True)

    st.caption("💡 Tip: Open the Excel file in your thesis appendix or paste the CSV into your research data.")


# ─────────────────────────────────────────────────────────────────────────────
# MODE: REVIEW UPLOADED FILE
# ─────────────────────────────────────────────────────────────────────────────
elif "Review" in mode:
    st.subheader("📄 Upload your mix design document")
    st.caption("Upload a PDF report, mix design sheet, or photo of a mix ticket. Claude will review it and generate a structured report.")

    with st.expander("📋 Project information (optional)"):
        pi1,pi2 = st.columns(2)
        with pi1:
            proj_name   = st.text_input("Project name", placeholder="e.g. RC4 Bridge Deck")
            location    = st.text_input("Location", placeholder="e.g. Brookings, SD")
        with pi2:
            prepared_by = st.text_input("Prepared by", placeholder="e.g. Dr. Mohammad Teymouri, PE")
            company     = st.text_input("Organization", placeholder="e.g. SDSU Teymouri Research Lab")
        cement_prod = flyash_prod = slag_prod = sf_prod = pcc_prod = ""

    uploaded = st.file_uploader("Upload mix design file",
                                type=["pdf","png","jpg","jpeg","webp"],
                                help="PDF, image of mix ticket, or lab report")
    review_notes = st.text_area("Additional context (optional)",
                                placeholder="e.g. Bridge deck, F2 exposure. Check ACI 318 w/cm compliance.",
                                height=80)

    if uploaded and st.button("🔍 Review Mix Design", use_container_width=True, type="primary"):
        with st.spinner("Analyzing your mix design document..."):
            file_bytes = uploaded.read()
            file_b64   = base64.b64encode(file_bytes).decode()
            is_pdf     = uploaded.type == "application/pdf"
            media_type = uploaded.type

            if demo_mode:
                review_text = """## Mix Design Summary
This is a DEMO analysis. Turn off Demo Mode and add your API key to get Claude to read the actual values from your document.

In live mode, Claude extracts: mix ID, w/cm, cement/SCM content, water, aggregate info, admixtures, density, target strength, and casting info.

## ACI 211.1 & ACI 318 Compliance Check
- Live mode performs a real compliance check against actual values in your document

## Recommended QC Tests
- ASTM C143 — Slump at discharge and placement
- ASTM C231 — Air content
- ASTM C39 — Compressive strength at 7 and 28 days
- ASTM C138 — Unit weight and yield"""
            else:
                REVIEW_PROMPT = f"""You are a concrete materials engineer reviewing a mix design sheet.
Carefully read all values in this document/image and provide a structured review:

## Mix Design Summary
Extract all key values: w/cm, cement content, SCM types and %, water, aggregate info, admixtures, density, target strength, mix ID, date.

## ACI 211.1 & ACI 318 Compliance Check
For each item state PASS / WARN / FAIL:
- w/cm vs exposure class requirements
- Minimum cementitious content
- Air content if shown
- Any other compliance items visible

## Observed Materials
List all materials with quantities from the document.

## Durability Flags
List concerns based on what you see.

## Recommended QC Tests
List appropriate ASTM tests.

## Summary & Recommendations
2-3 sentences of practical guidance.

Additional context: {review_notes or 'None provided'}
Be specific — use actual numbers from the document."""

                api_key = os.getenv("ANTHROPIC_API_KEY","")
                client  = anthropic.Anthropic(api_key=api_key)
                content_msg = [
                    {"type":"document","source":{"type":"base64","media_type":"application/pdf","data":file_b64}}
                    if is_pdf else
                    {"type":"image","source":{"type":"base64","media_type":media_type,"data":file_b64}},
                    {"type":"text","text":REVIEW_PROMPT}
                ]
                msg = client.messages.create(model="claude-opus-4-5", max_tokens=2000,
                                             messages=[{"role":"user","content":content_msg}])
                review_text = msg.content[0].text

        st.subheader("Review findings")
        st.markdown(review_text)

        pi = ProjectInfo(project_name=proj_name, location=location,
                         prepared_by=prepared_by, company=company,
                         cement_producer=cement_prod, flyash_producer=flyash_prod,
                         slag_producer=slag_prod, sf_producer=sf_prod, pcc_producer=pcc_prod)
        dummy_inp = MixDesignInput(
            fc_psi=0, exposure_codes=["—"], agg_size="—", slump="—",
            fm_fa=0, sg_ca=0, sg_fa=0, rodded_density_ca=0, agg_shape="—",
            field_notes=review_notes, project_info=pi, uploaded_file_name=uploaded.name)
        dummy_aci = {
            "design_fc_psi":0,"selected_wcm":0,"air_pct":0,"density_pcf":0,"bv_ca":0,
            "proportions":{k:0 for k in ["water_lbs","cement_lbs","flyash_lbs","slag_lbs","sf_lbs","pcc_lbs","ca_lbs","fa_lbs","total_cm_lbs"]},
            "volumes_ft3":{k:0 for k in ["water","cement","flyash","slag","sf","pcc","ca","fa","air","total"]},
            "flags":[{"flag":"See review findings above.","status":"ok"}],
        }
        review_result = MixDesignResult(input_summary=dummy_inp, aci_result=dummy_aci,
            risk_level="Review", ai_analysis="Document reviewed.",
            recommendations=review_text, file_review_notes=review_text)

        st.divider()
        pdf_bytes = generate_pdf_report(review_result)
        st.download_button("📄 Download Review Report (PDF)", pdf_bytes,
                           f"review_{uploaded.name}.pdf", "application/pdf",
                           use_container_width=True)


# ─────────────────────────────────────────────────────────────────────────────
# MODE: STUDENT Q&A
# ─────────────────────────────────────────────────────────────────────────────
elif "Q&A" in mode:
    st.markdown("""
    <div class="sdsu-header">💬 &nbsp; Concrete Q&A — Graduate Research Assistant</div>
    <div class="sdsu-gold-bar"></div>
    """, unsafe_allow_html=True)
    st.caption("Ask about ASTM/ACI standards, concrete properties, mix design theory, durability, SCMs, UHPC, SCC, or any concrete topic.")

    st.markdown('<div class="sdsu-q-label">Quick questions to get started</div>', unsafe_allow_html=True)
    q1,q2,q3 = st.columns(3)
    suggestions = [
        ("What is ASTM C143 — slump test?", q1),
        ("What is ASTM C39 — compressive strength test?", q2),
        ("What is ASTM C231 — air content test?", q3),
        ("What is ASTM C1202 — RCPT test?", q1),
        ("What are SCC tests and standards?", q2),
        ("What is UHPC and how is it tested?", q3),
        ("How is crack detection done in concrete?", q1),
        ("What is the Vebe test for pavement concrete?", q2),
        ("What is fly ash Class C vs Class F?", q3),
        ("What are SCMs in concrete?", q1),
        ("What is the ITZ in concrete microstructure?", q2),
        ("How does carbonation affect concrete durability?", q3),
    ]
    for question, col in suggestions:
        with col:
            if st.button(question, use_container_width=True, key=f"q_{question[:20]}"):
                st.session_state.chat_input_prefill = question

    st.divider()

    if "chat_history" not in st.session_state:
        st.session_state.chat_history = []

    CONCRETE_SYSTEM = """You are a concrete materials engineering assistant and tutor at the Teymouri Research Lab,
South Dakota State University (SDSU). You help construction engineering graduate students with research and thesis work.

When asked about a specific ASTM or ACI standard, structure your answer as:
1. Standard designation (full, e.g. ASTM C143-20)
2. Purpose — one sentence
3. Equipment required
4. Step-by-step procedure — numbered, concise
5. Key acceptance criteria / limits
6. Google search terms to find the free standard
7. Related standards

For graduate-level topics (ITZ, ASR, carbonation, chloride diffusion, pozzolanic reaction), give in-depth mechanistic explanations.

Key standards: C143 (slump), C39 (compressive), C231 (air content), C138 (unit weight),
C1202 (RCPT), C666 (freeze-thaw), C157 (shrinkage), C618 (fly ash), C989 (slag),
C1240 (silica fume), C494 (admixtures), C172 (sampling), C192 (making specimens),
C1611 (SCC slump flow), C1621 (J-ring), C1170 (Vebe/RCC).

PCC research context (Teymouri Research Lab): Precipitated Calcium Carbonate from Western Sugar Cooperative
is studied as an inert micro-filler. At 3-8% replacement, particle packing improves microstructure
and provides 5-9% compressive strength gain. No pozzolanic activity. SG ≈ 2.71.
Cite this as: Teymouri Research Lab, SDSU (ongoing) when relevant.

For SCMs question: explain Portland cement, fly ash (C/F), slag, silica fume, PCC.
For ITZ: explain interfacial transition zone microstructure and effect on strength.
For carbonation: CO2 + Ca(OH)2 → CaCO3, drops pH, depassivates rebar.
For ASR: alkali + reactive silica → gel, absorbs water, expands, cracks.

Always be rigorous and cite ACI/ASTM standard numbers. Encourage graduate-level thinking."""

    DEMO_RESPONSES = {
        "c143": """**ASTM C143 — Standard Test Method for Slump of Hydraulic-Cement Concrete**

**Purpose:** Measures the workability (consistency) of fresh concrete before placement.

**Equipment:** Abrams slump cone (4\" top, 8\" base, 12\" height), tamping rod (5/8\" × 24\", bullet-tipped), rigid base plate, ruler.

**Step-by-step procedure:**
1. Dampen cone and base plate; hold cone steady by standing on foot pieces
2. Sample concrete per **ASTM C172** — begin test within 5 minutes
3. Fill in **3 equal layers** (~4\" each); rod each layer **25 times** uniformly
4. For upper layers, penetrate 1\" into the previous layer with each rod stroke
5. Strike off top flush; lift cone **straight up in 5 ± 2 seconds** — no twisting
6. Measure drop from cone height to displaced top of concrete to nearest **1/4\"**
7. Complete test within **2.5 minutes** of sampling

**Target slumps (ACI 211.1):**
| Construction type | Slump |
|---|---|
| Footings / substructure walls | 1–3\" |
| Beams / reinforced walls | 1–4\" |
| Pavements / slabs | 1–3\" |
| Pumped concrete | 4–6\" |

**Invalid test:** Shear slump (concrete falls apart sideways) → discard, resample, retest.

**Find it — search Google for:**
- 🔍 `ASTM C143 slump test procedure PDF`
- 🔍 `concrete slump test Abrams cone step by step`""",

        "c39": """**ASTM C39 — Compressive Strength of Cylindrical Concrete Specimens**

**Purpose:** Measures f'c — the primary acceptance and quality control test for concrete.

**Equipment:** Compression testing machine (calibrated per ASTM E4), 4×8\" or 6×12\" cylinder molds, neoprene pads (ASTM C1231) or sulfur capping compound.

**Step-by-step:**
1. Make specimens per **ASTM C192** (lab) or **ASTM C31** (field) — fill in 2–3 layers, rod 25× each
2. Cure in molds at 60–80°F for **24 ± 8 hours**
3. Strip molds; moist cure at **73 ± 3°F** until test age
4. Prepare ends: cap or grind flat within **0.002\"**
5. Center specimen; apply load at **35 ± 7 psi/second** continuously
6. Record maximum load; calculate **f'c = Load ÷ Area**
7. Note fracture pattern (Type 1–6 per Figure 3)

**Test ages:** 7 days (early check), **28 days** (acceptance)
**Acceptance (ACI 318):** Average of 2 cylinders ≥ f'c; no single cylinder < f'c − 500 psi

**Find it — search Google for:**
- 🔍 `ASTM C39 compressive strength concrete cylinder PDF`
- 🔍 `concrete cylinder test 28 day procedure ACI`""",

        "c231": """**ASTM C231 — Air Content by Pressure Method**

**Purpose:** Measures total air content (%) in fresh normal-weight concrete using Boyle's Law.

**Equipment:** Type A or B pressure air meter, tamping rod, vibrator, rubber mallet.

**Step-by-step:**
1. Sample per ASTM C172; fill meter bowl in **3 equal layers**, rod 25× each + 10–15 mallet taps
2. Strike off top; clean flange; clamp cover assembly
3. Inject water through petcocks to remove air above concrete; close petcocks
4. Pump to initial pressure; tap sides; open main valve; read gauge when stable
5. Record to nearest **0.1%**
6. Subtract **aggregate correction factor** (determined per Annex A2)

**Required air content (3/4\" aggregate, ACI 318):**
| Exposure | Target air % |
|---|---|
| F1 — moderate freeze-thaw | 5.0% |
| F2 — severe freeze-thaw | 6.0% |
| F3 — deicers | 6.0% |

**Note:** For lightweight or slag aggregate, use **ASTM C173** (volumetric method) instead.

**Find it — search Google for:**
- 🔍 `ASTM C231 air content pressure meter concrete`
- 🔍 `concrete air entrainment test procedure freeze thaw`""",

        "rcpt": """**ASTM C1202 — Rapid Chloride Permeability Test (RCPT)**

**Purpose:** Measures electrical conductance through concrete as an indicator of resistance to chloride ion penetration — key durability parameter for bridge decks, marine structures, deicing environments.

**Equipment:** Two electrode cells, 60V DC power supply, ammeter/data logger, vacuum pump, 4\"×2\" disc specimens.

**Step-by-step:**
1. Cut **2\" thick disc** from 4×8\" cylinder at mid-height (wet saw)
2. Vacuum saturate in water for **18 hours** (Annex A)
3. Mount disc between cells: one side **3% NaCl**, other side **0.3N NaOH**
4. Apply **60V DC** for **6 hours**; record current every 30 min
5. Calculate total charge (coulombs) = area under current-time curve

**Permeability classification:**
| Coulombs | Rating |
|---|---|
| > 4,000 | High |
| 2,000–4,000 | Moderate |
| 1,000–2,000 | Low |
| 100–1,000 | Very low |
| < 100 | Negligible |

**Research note:** Test at **56 days** for SCM-rich mixes (fly ash, slag) — early testing underestimates durability because SCM reactions are incomplete at 28 days.
ASTM C1556 (bulk diffusion) is more accurate for high-SCM mixes.

**Find it — search Google for:**
- 🔍 `ASTM C1202 rapid chloride permeability test PDF`
- 🔍 `RCPT concrete durability coulombs bridge deck`""",

        "scc": """**Self-Consolidating Concrete (SCC) — Key Tests (ASTM C1611, C1621, C1610)**

**What is SCC?** Flows under its own weight, fully fills formwork, passes reinforcement — no vibration needed. Standard slump test is insufficient; requires these 4 tests:

**1. ASTM C1611 — Slump Flow + T50**
- Invert slump cone on flat plate; lift; measure spread diameter
- **Target:** 18–32\" (450–810 mm)
- **T50:** Time to reach 20\" → measures viscosity (target: 2–7 sec)
- **VSI (Visual Stability Index):** 0–1 = stable, 2–3 = unstable

**2. ASTM C1621 — J-Ring (passing ability)**
- Slump flow WITH vertical rebar ring around cone
- **Max difference from free slump:** ≤ 2\" → good passing ability
- Larger difference = risk of blockage in congested reinforcement

**3. ASTM C1610 — Column Segregation**
- Fill 26\" PVC column; wait 15 min; section into 4 parts
- Compare aggregate top vs bottom
- **Segregation index ≤ 10%** = acceptable

**4. ASTM C1712 — Penetration Test (site)**
- Quick check: depth a cylinder penetrates fresh SCC

**Typical SCC mix:** w/cm 0.32–0.42 | paste volume 34–40% | HRWRA (ASTM C494 Type F/G)

**Find it — search Google for:**
- 🔍 `ASTM C1611 slump flow SCC test procedure`
- 🔍 `ACI 237R self consolidating concrete guide`""",

        "uhpc": """**Ultra-High Performance Concrete (UHPC) — Properties & Testing**

**Definition:** f'c ≥ 14,500 psi (100 MPa), exceptional durability, high tensile strength from steel fibers. No coarse aggregate.

**Typical UHPC composition (per CY):**
| Material | Amount |
|---|---|
| Portland cement | 700–900 lb |
| Silica fume | 200–250 lb (25–30%) |
| Quartz flour (< 300 μm) | 200–400 lb |
| Quartz sand (150–600 μm) | 900–1100 lb |
| Steel fibers (2% vol.) | ~260 lb |
| HRWRA | 30–50 lb |
| w/cm | 0.14–0.22 |

**Key test standards:**
| Test | Standard | UHPC typical |
|---|---|---|
| Compressive strength | ASTM C39 | 14,500–29,000 psi |
| Flexural performance | **ASTM C1609** | 2,000–4,000 psi |
| Splitting tensile | ASTM C496 | 1,000–2,500 psi |
| Flow (fresh) | ASTM C1437 | > 8\" spread |
| Chloride permeability | ASTM C1202 | < 100 coulombs |

**ASTM C1609** (most important for UHPC): 6×6×20\" beam, 18\" span → full load-deflection curve → captures post-crack toughness from fiber bridging.

**Curing:** Steam at **194°F (90°C) for 48 hours** to achieve full matrix densification.

**Find it — search Google for:**
- 🔍 `FHWA HRT-14-084 UHPC state of the art report free PDF`
- 🔍 `ASTM C1609 fiber reinforced concrete flexural test`""",

        "crack": """**Crack Detection in Concrete — NDT Methods & Standards**

**Visual inspection (always first):**
- Crack comparator gauge → measure width to nearest 0.001\"
- Map type: map/shrinkage, structural flexural, corrosion-induced, settlement
- ACI 224R limits: > 0.013\" (0.33 mm) = concern for exterior exposure

**NDT methods:**

**1. ASTM C597 — Ultrasonic Pulse Velocity (UPV)**
- Transmit ultrasonic pulse; measure travel time
- Sound concrete: > 14,000 ft/s (4,300 m/s)
- Crack depth: indirect transmission method

**2. ASTM C1383 — Impact-Echo**
- Strike surface; measure reflected stress wave frequency
- Detects delamination, voids, internal cracks in slabs

**3. Ground Penetrating Radar (ASTM D6432)**
- Electromagnetic pulse maps cracks, rebar, voids
- Fast scan method for bridge decks

**4. Acoustic Emission (ASTM E1316)**
- Passive sensors detect waves from active crack propagation
- Real-time structural monitoring

**Crack width guide (ACI 224R):**
| Width | Action |
|---|---|
| < 0.006\" | Hairline — monitor |
| 0.006–0.013\" | Fine — acceptable outdoors |
| > 0.013\" | Wide — investigate |
| > 0.020\" | Severe — structural concern |

**Find it — search Google for:**
- 🔍 `ACI 224R crack control concrete structures PDF`
- 🔍 `concrete NDT crack detection impact echo UPV`""",

        "vebe": """**Vebe Test — ASTM C1170 (Stiff / Roller-Compacted Concrete Pavement)**

**Why not slump?** Pavement and RCC mixes are intentionally stiff (0\" slump). The slump test cannot differentiate between these mixes — Vebe time fills this gap.

**Equipment:** Vibrating table (50 Hz, 0.35 mm amplitude), Abrams slump cone, cylindrical container (9.5\" dia × 8\" H), Plexiglass rider disc, stopwatch.

**Step-by-step:**
1. Place slump cone inside cylindrical container on vibrating table
2. Fill in 3 layers, rod 25× each (identical to slump test)
3. Strike off top; remove slump cone
4. Lower Plexiglass disc onto concrete surface
5. **Start vibrator and stopwatch simultaneously**
6. Stop when disc underside is **uniformly covered with mortar** (no escaping air bubbles)
7. Record time to nearest **0.5 second** = Vebe time

**Vebe time classification:**
| Vebe time | Class | Typical use |
|---|---|---|
| 3–5 s | V0 | Wet RCC |
| 6–12 s | V1 | Standard RCC pavement |
| 12–20 s | V2 | Dry RCC, lean concrete base |
| 20–30 s | V3 | Very stiff pavement |
| > 30 s | V4 | Extremely stiff, dry-cast |

**RCC pavement target:** **6–12 seconds (V1)** for optimal vibratory roller compaction.

**Find it — search Google for:**
- 🔍 `ASTM C1170 Vebe test RCC roller compacted concrete`
- 🔍 `Vebe consistometer pavement concrete workability`""",

        "fly ash": """**Fly Ash in Concrete — ASTM C618 (Class C vs Class F)**

**What is fly ash?** Fine powder byproduct of coal combustion, captured by electrostatic precipitators. Partial cement replacement.

**Class F vs Class C:**
| Property | Class F | Class C |
|---|---|---|
| Coal source | Anthracite/bituminous | Sub-bituminous/lignite |
| SiO₂+Al₂O₃+Fe₂O₃ | > 70% | > 50% |
| CaO content | Low (< 10%) | High (15–35%) |
| Reactivity | Pozzolanic only | Pozzolanic + cementitious |
| Self-cementing? | No | Yes |
| Strength gain | Slower, long-term | Faster |
| Sulfate resistance | Better | Good |
| Midwest availability | Less common | **Very common in SD** |

**Pozzolanic reaction:** SiO₂ (fly ash) + Ca(OH)₂ (cement hydration) → C-S-H (additional binding gel)
This reaction is slow — benefits appear at **28–90 days**.

**ASTM C618 requirements:**
- Fineness: ≤ 34% retained on No. 325 sieve
- Strength activity index: ≥ 75% at 28 days
- Soundness: autoclave expansion ≤ 0.8%

**Find it — search Google for:**
- 🔍 `ASTM C618 fly ash concrete specification`
- 🔍 `Class C vs Class F fly ash concrete difference`""",

        "scm_what": """**Supplementary Cementitious Materials (SCMs) in Concrete**

**Definition:** Materials that partially replace Portland cement. They react with cement hydration products to form additional C-S-H (calcium silicate hydrate) — the main binding compound in concrete.

**The 5 main SCMs:**

| SCM | Standard | Reactivity | Typical % | Key benefit |
|---|---|---|---|---|
| **Fly ash (Class F)** | ASTM C618 | Pozzolanic | 15–25% | Durability, cost |
| **Fly ash (Class C)** | ASTM C618 | Pozzolanic + cementitious | 20–35% | Faster than Class F |
| **Slag cement** | ASTM C989 | Latent hydraulic | 25–50% | Low permeability |
| **Silica fume** | ASTM C1240 | Pozzolanic (high reactivity) | 5–10% | Very high strength |
| **PCC** (Teymouri Lab) | — | **Inert filler** | 3–8% | Particle packing |

**Pozzolanic reaction:**
> SiO₂ (SCM) + Ca(OH)₂ (from cement) + H₂O → **C-S-H** (additional binding gel)

**Why use SCMs?**
- Reduce Portland cement → lower CO₂ (cement production = 8% of global CO₂)
- Improve long-term strength and durability
- Reduce heat of hydration (mass concrete)
- Lower permeability and chloride diffusion
- Cost savings

**PCC exception — Teymouri Research Lab:** Precipitated Calcium Carbonate from Western Sugar Cooperative is an **inert micro-filler** — no pozzolanic reaction. Works through physical particle packing. At 5% replacement: +5–9% compressive strength gain from improved packing density.

**Find it — search Google for:**
- 🔍 `supplementary cementitious materials SCM concrete ACI`
- 🔍 `fly ash slag silica fume concrete SCM comparison`""",

        "itz": """**Interfacial Transition Zone (ITZ) in Concrete Microstructure**

**What is the ITZ?** The ~20–50 μm region of cement paste surrounding aggregate particles. It is the **weakest zone in concrete** — critical for understanding strength and durability.

**Why does ITZ exist?**
During mixing, water films form around aggregate surfaces (wall effect). This creates a zone of:
- Higher water-cement ratio than bulk paste
- More porous microstructure
- Larger Ca(OH)₂ crystals (hexagonal plates, oriented perpendicular to aggregate)
- Less C-S-H content

**ITZ microstructure:**
- 0–5 μm from aggregate: mostly Ca(OH)₂ duplex film
- 5–20 μm: ettringite + Ca(OH)₂ rich zone
- 20–50 μm: gradual transition to bulk paste
- > 50 μm: bulk cement paste (normal density)

**Effect on concrete properties:**
- **Strength:** Cracks initiate at ITZ under load → limits compressive and tensile strength
- **Permeability:** ITZ provides preferential pathway for water and chloride ingress
- **Durability:** Weakest link for freeze-thaw, sulfate attack, ASR

**How SCMs improve ITZ:**
- Silica fume (finest SCM, ~0.1 μm): fills voids in ITZ → dramatically improves bond
- Fly ash and slag: react with Ca(OH)₂ → convert to C-S-H → densify ITZ
- Lower w/cm → less wall effect → thinner, denser ITZ

**Graduate research note:** ASTM C457 (air void analysis) and SEM/EDS imaging are used to characterize ITZ microstructure in research.

**Find it — search Google for:**
- 🔍 `interfacial transition zone ITZ concrete microstructure`
- 🔍 `ITZ concrete SEM imaging aggregate paste bond`""",

        "carbonation": """**Carbonation of Concrete — Mechanism & Durability Impact**

**What is carbonation?**
CO₂ from the atmosphere diffuses into concrete and reacts with hydration products:

> CO₂ + Ca(OH)₂ → **CaCO₃** + H₂O
> CO₂ + C-S-H → CaCO₃ + amorphous silica gel

**Why it matters — pH drop:**
- Fresh concrete pH: **12.5–13.5** → passive oxide film protects rebar
- After carbonation: pH drops to **~8–9** → passive film destroyed → **depassivation**
- Result: rebar corrodes freely → expansive corrosion products → cracking and spalling

**Carbonation depth over time:**
> d = K × √t
- d = carbonation depth (mm)
- K = carbonation rate coefficient (depends on w/cm, cement type, CO₂ exposure)
- t = time (years)

**Factors that increase carbonation rate:**
| Factor | Effect |
|---|---|
| Higher w/cm | More permeable → faster CO₂ diffusion |
| SCM-rich mixes | Less Ca(OH)₂ → carbonation front advances faster |
| Poor curing | More permeable surface zone |
| Low CO₂ environments | Slower (typical outdoor: 0.04% CO₂) |

**Testing carbonation depth — ASTM / RILEM:**
- Core or cut section → spray with **phenolphthalein indicator**
- Uncarbonated zone (pH > 9): **pink/magenta**
- Carbonated zone (pH < 9): **colorless**
- Measure depth to color change

**Protective measures:**
- Low w/cm (≤ 0.45), adequate curing, proper cover depth (ACI 318 Table 20.6.1)
- Carbonation-resistant coatings for exposed elements

**Find it — search Google for:**
- 🔍 `concrete carbonation mechanism rebar corrosion`
- 🔍 `phenolphthalein carbonation depth test concrete`"""
    }

    def get_demo_response(question: str) -> str:
        q = question.lower()
        if "c1202" in q or "rcpt" in q or "rapid chloride" in q: return DEMO_RESPONSES["rcpt"]
        elif "scc" in q or "self-consolidat" in q or "j-ring" in q or "slump flow" in q: return DEMO_RESPONSES["scc"]
        elif "uhpc" in q or "ultra-high" in q or "ultra high" in q: return DEMO_RESPONSES["uhpc"]
        elif "crack" in q and ("detect" in q or "test" in q or "measur" in q or "done" in q): return DEMO_RESPONSES["crack"]
        elif "vebe" in q or ("pavement" in q and ("rcc" in q or "roller" in q or "test" in q)): return DEMO_RESPONSES["vebe"]
        elif "c143" in q or ("slump" in q and "vebe" not in q and "flow" not in q): return DEMO_RESPONSES["c143"]
        elif "c39" in q or "compressive" in q: return DEMO_RESPONSES["c39"]
        elif "c231" in q or "air content" in q: return DEMO_RESPONSES["c231"]
        elif "fly ash" in q or "class c" in q or "class f" in q: return DEMO_RESPONSES["fly ash"]
        elif "scm" in q or "supplementary" in q or "cementitious material" in q: return DEMO_RESPONSES["scm_what"]
        elif "itz" in q or "interfacial" in q or "transition zone" in q: return DEMO_RESPONSES["itz"]
        elif "carbonat" in q: return DEMO_RESPONSES["carbonation"]
        else:
            return ("**Demo mode — this specific question is not pre-loaded.**\n\n"
                    "Turn off Demo Mode and add your API key to get a full answer.\n\n"
                    "**Pre-loaded topics:**\n"
                    "ASTM C143 · C39 · C231 · C1202 (RCPT) · SCC · UHPC · "
                    "Crack detection · Vebe test · Fly ash C vs F · SCMs · ITZ · Carbonation")

    for msg in st.session_state.chat_history:
        with st.chat_message(msg["role"]):
            st.markdown(msg["content"])

    prefill    = st.session_state.pop("chat_input_prefill", "")
    user_input = st.chat_input("Ask a concrete question...", key="concrete_chat")
    active_input = prefill or user_input

    if active_input:
        st.session_state.chat_history.append({"role":"user","content":active_input})
        with st.chat_message("user"):
            st.markdown(active_input)

        with st.chat_message("assistant"):
            with st.spinner("Looking that up..."):
                if demo_mode:
                    response = get_demo_response(active_input)
                else:
                    api_key = os.getenv("ANTHROPIC_API_KEY","")
                    client  = anthropic.Anthropic(api_key=api_key)
                    messages = [{"role":m["role"],"content":m["content"]} for m in st.session_state.chat_history]
                    resp = client.messages.create(model="claude-opus-4-5", max_tokens=1500,
                                                  system=CONCRETE_SYSTEM, messages=messages)
                    response = resp.content[0].text
            st.markdown(response)
            st.session_state.chat_history.append({"role":"assistant","content":response})

    if st.session_state.get("chat_history"):
        st.markdown("<div style='height:3px;background:linear-gradient(to right,#0033A0,#FFB71B);border-radius:2px;margin:12px 0;'></div>", unsafe_allow_html=True)
        col_c, _ = st.columns([1,4])
        with col_c:
            if st.button("🗑️ Clear chat", use_container_width=True):
                st.session_state.chat_history = []
                st.rerun()
